import argparse
import os
import json
import pdfplumber
import google.generativeai as genai
from openpyxl import load_workbook

def main():
    parser = argparse.ArgumentParser(description="Process PDFs, extract data using LLM, and populate an Excel file.")
    parser.add_argument(
        "--pdf-dir",
        type=str,
        default=".",
        help="Directory containing PDF files to process. Defaults to the current directory."
    )
    parser.add_argument(
        "--template",
        type=str,
        default="base.exel",
        help="Path to the Excel template file. Defaults to 'base.exel' in the current directory."
    )
    parser.add_argument(
        "--output",
        type=str,
        default="output.xlsx",
        help="Path for the generated output Excel file. Defaults to 'output.xlsx'."
    )
    parser.add_argument(
        "--prompt-file",
        type=str,
        default="prompt.txt",
        help="Path to the prompt file for the LLM. Defaults to 'prompt.txt'."
    )
    parser.add_argument(
        "--gemini-api-key",
        type=str,
        required=True,
        help="Google Gemini API key. Can also be set via the GOOGLE_API_KEY environment variable."
    )

    args = parser.parse_args()

    
    print(f"PDF Directory: {args.pdf_dir}")
    print(f"Template File: {args.template}")
    print(f"Output File: {args.output}")
    print(f"Prompt File: {args.prompt_file}")
    print(f"Gemini API Key: {'*' * len(args.gemini_api_key)}")

    
    if not os.path.isdir(args.pdf_dir):
        print(f"Error: PDF directory not found: {args.pdf_dir}")
        return

    if not os.path.isfile(args.template):
        print(f"Error: Template file not found: {args.template}")
        return

    if not os.path.isfile(args.prompt_file):
        print(f"Error: Prompt file not found: {args.prompt_file}")
        return

    
    try:
        with open(args.prompt_file, 'r', encoding='utf-8') as f:
            base_prompt = f.read()
        print("Prompt loaded successfully.")
    except IOError as e:
        print(f"Error reading prompt file {args.prompt_file}: {e}")
        return

    
    try:
        genai.configure(api_key=args.gemini_api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        print("Gemini API configured.")
    except Exception as e:
        print(f"Error configuring Gemini API: {e}")
        return

    
    try:
        wb = load_workbook(args.template)
        sheet = wb["Cab-Usage"] 
        print("Excel template loaded.")
    except FileNotFoundError:
        print(f"Error: Template file not found at {args.template}")
        return
    except KeyError:
        print("Error: Sheet 'Cab-Usage' not found in the template Excel file.")
        return
    except Exception as e:
        print(f"Error loading Excel template: {e}")
        return

    
    pdf_files = [f for f in os.listdir(args.pdf_dir) if f.lower().endswith(".pdf")]
    if not pdf_files:
        print(f"No PDF files found in directory: {args.pdf_dir}")
        return
    
    print(f"Found {len(pdf_files)} PDF file(s) to process.")

    
    START_ROW = 9
    COLUMN_MAPPING = {
        "idx": 0, 
        "date": 1, 
        "inward_from": 2, 
        "inward_to": 3, 
        "outward_from": 4, 
        "outward_to": 5, 
        "visits": 6, 
        "vendor": 7, 
        "inward_charges": 8, 
        "outward_charges": 9 
    }

    all_trip_data = []

    
    for i, pdf_filename in enumerate(pdf_files):
        pdf_path = os.path.join(args.pdf_dir, pdf_filename)
        print(f"\nProcessing {pdf_filename} ({i+1}/{len(pdf_files)})...")

        
        try:
            extracted_text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    extracted_text += page.extract_text() + "\n" if page.extract_text() else ""
            if not extracted_text.strip():
                print(f"Warning: No text could be extracted from {pdf_filename}. Skipping.")
                continue
            print(f"Text extracted from {pdf_filename}.")
        except Exception as e:
            print(f"Error extracting text from {pdf_filename}: {e}. Skipping this file.")
            continue

        
        full_prompt = base_prompt.replace("[PDF TEXT WILL BE INSERTED HERE]", extracted_text.strip())
        try:
            print(f"Sending request to Gemini for {pdf_filename}...")
            response = model.generate_content(
                full_prompt,
                generation_config=genai.types.GenerationConfig(
                    response_mime_type="application/json",
                )
            )
            if not response.candidates or not response.candidates[0].content.parts:
                print(f"Error: No valid response from Gemini for {pdf_filename}. Skipping.")
                continue
            json_response_str = response.candidates[0].content.parts[0].text
            print(f"Received response from Gemini for {pdf_filename}.")
        except Exception as e:
            print(f"Error calling Gemini API for {pdf_filename}: {e}. Skipping this file.")
            continue

        
        try:
            data = json.loads(json_response_str)
            
            for key in ["date", "inward_from", "inward_to", "outward_from", "outward_to", "visits", "vendor", "inward_charges", "outward_charges"]:
                if key not in data:
                    data[key] = "" 
            if not data.get("date"):
                print(f"Warning: No date found in response for {pdf_filename}. Skipping this entry.")
                continue
            print("JSON parsed successfully.")
            all_trip_data.append(data)
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response for {pdf_filename}: {e}. Response was: {json_response_str}. Skipping this file.")
            continue

    if not all_trip_data:
        print("\nNo valid trip data extracted from any PDF. Exiting.")
        return

    
    date_to_row_map = {}
    next_available_row = START_ROW
    
    excel_rows_data = {}

    print("\nConsolidating data by date...")
    for item_idx, trip_item in enumerate(all_trip_data):
        date = trip_item.get("date")
        if not date:
            continue 

        if date in date_to_row_map:
            
            row_num = date_to_row_map[date]
            existing_row_data = excel_rows_data.get(row_num, {})
            
            print(f"Updating existing row {row_num} for date {date} with new trip data (from item {item_idx +1}).")

            
            
            for key in ["inward_from", "inward_to", "inward_charges", "outward_from", "outward_to", "outward_charges", "vendor", "visits"]:
                current_val = existing_row_data.get(key, "")
                new_val = trip_item.get(key, "")
                if not current_val and new_val: 
                    existing_row_data[key] = new_val
            excel_rows_data[row_num] = existing_row_data

        else:
            
            row_num = next_available_row
            date_to_row_map[date] = row_num
            
            
            new_row_data = {
                "idx": len(excel_rows_data) + 1, 
                "date": date,
                "inward_from": trip_item.get("inward_from", ""),
                "inward_to": trip_item.get("inward_to", ""),
                "outward_from": trip_item.get("outward_from", ""),
                "outward_to": trip_item.get("outward_to", ""),
                "visits": trip_item.get("visits", 1),
                "vendor": trip_item.get("vendor", ""),
                "inward_charges": str(trip_item.get("inward_charges", "")),
                "outward_charges": str(trip_item.get("outward_charges", ""))
            }
            excel_rows_data[row_num] = new_row_data
            print(f"Assigned new row {row_num} for date {date} (from item {item_idx +1}).")
            next_available_row += 1
            
    
    
    
    sorted_excel_items = sorted(excel_rows_data.items(), key=lambda item: item[1]['date'])
    
    
    final_excel_data = {}
    current_excel_row = START_ROW
    for _, row_data in sorted_excel_items:
        final_excel_data[current_excel_row] = row_data
        
        final_excel_data[current_excel_row]['idx'] = current_excel_row - START_ROW + 1
        current_excel_row += 1

    
    print("\nPopulating Excel sheet with sorted, consolidated data...")
    if final_excel_data:
        try:
            for row_num, row_data in final_excel_data.items():
                sheet.cell(row=row_num, column=COLUMN_MAPPING["idx"] + 1, value=row_data.get("idx"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["date"] + 1, value=row_data.get("date"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_from"] + 1, value=row_data.get("inward_from"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_to"] + 1, value=row_data.get("inward_to"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_from"] + 1, value=row_data.get("outward_from"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_to"] + 1, value=row_data.get("outward_to"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["visits"] + 1, value=row_data.get("visits"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["vendor"] + 1, value=row_data.get("vendor"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_charges"] + 1, value=row_data.get("inward_charges"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_charges"] + 1, value=row_data.get("outward_charges"))
            
            
            wb.save(args.output)
            print(f"\nSuccessfully processed. Consolidated and sorted data for {len(final_excel_data)} unique date(s) into Excel. Output saved to {args.output}")
        except Exception as e:
            print(f"\nError populating or saving Excel file: {e}")
    else:
        print("\nNo data was consolidated to write to the Excel file.")


