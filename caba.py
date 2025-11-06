"""
CABA - Cab Booking Analyzer CLI
A professional command-line interface for processing PDF cab booking data and generating Excel reports.
"""

import os
import json
import argparse
from pathlib import Path
from typing import Optional, Dict, Any
import pdfplumber
import google.generativeai as genai
from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.text import Text
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn
from rich.prompt import Prompt, Confirm
from rich.style import Style
from rich.theme import Theme
import click

custom_theme = Theme({
    "primary": "bold cyan",
    "secondary": "bold magenta",
    "success": "bold green",
    "warning": "bold yellow",
    "error": "bold red",
    "info": "bold blue",
    "highlight": "bold white on blue"
})

console = Console(theme=custom_theme)

class CabaConfig:
    """Handles configuration management for CABA"""
    
    def __init__(self):
        self.config_file = "config.json"
        self.config = self.load_config()
    
    def load_config(self) -> Dict[str, Any]:
        """Load configuration from file"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError):
                return {}
        return {}
    
    def save_config(self):
        """Save configuration to file"""
        try:
            with open(self.config_file, 'w') as f:
                json.dump(self.config, f, indent=2)
        except IOError:
            console.print("[warning]Warning: Could not save configuration[/warning]")
    
    def get(self, key: str, default: Any = None) -> Any:
        """Get configuration value"""
        return self.config.get(key, default)
    
    def set(self, key: str, value: Any):
        """Set configuration value"""
        self.config[key] = value
        self.save_config()

class PDFProcessor:
    """Handles PDF processing and Excel generation"""
    
    def __init__(self, gemini_api_key: str):
        self.gemini_api_key = gemini_api_key
        try:
            genai.configure(api_key=gemini_api_key)
            self.model = genai.GenerativeModel('gemini-2.5-flash')
        except Exception as e:
            console.print(f"[error]Error configuring Gemini API: {e}[/error]")
            raise
    
    def load_prompt(self, prompt_file: str) -> str:
        """Load prompt from file"""
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except IOError as e:
            console.print(f"[error]Error reading prompt file {prompt_file}: {e}[/error]")
            raise
    
    def process_pdfs(self, pdf_dir: str, template_file: str, output_file: str, 
                    prompt_file: str, progress_callback=None) -> bool:
        """Process all PDFs and generate Excel file"""
        
        try:
            base_prompt = self.load_prompt(prompt_file)
        except Exception:
            return False
        
        try:
            wb = load_workbook(template_file)
            sheet = wb["Cab-Usage"]
        except FileNotFoundError:
            console.print(f"[error]Template file not found: {template_file}[/error]")
            return False
        except KeyError:
            console.print("[error]Sheet 'Cab-Usage' not found in template[/error]")
            return False
        
        pdf_files = [f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")]
        if not pdf_files:
            console.print(f"[warning]No PDF files found in {pdf_dir}[/warning]")
            return False
        
        console.print(f"[info]Found {len(pdf_files)} PDF(s) to process[/info]")
        
        START_ROW = 9
        COLUMN_MAPPING = {
            "idx": 0, "date": 1, "inward_from": 2, "inward_to": 3,
            "outward_from": 4, "outward_to": 5, "visits": 6,
            "vendor": 7, "inward_charges": 8, "outward_charges": 9
        }
        
        all_trip_data = []
        
        for i, pdf_filename in enumerate(pdf_files):
            if progress_callback:
                progress_callback(i + 1, len(pdf_files), pdf_filename)
            
            pdf_path = os.path.join(pdf_dir, pdf_filename)
            
            try:
                extracted_text = ""
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        extracted_text += page.extract_text() + "\n" if page.extract_text() else ""
                
                if not extracted_text.strip():
                    console.print(f"[warning]No text extracted from {pdf_filename}[/warning]")
                    continue
            except Exception as e:
                console.print(f"[warning]Error extracting text from {pdf_filename}: {e}[/warning]")
                continue
            
            full_prompt = base_prompt.replace("[PDF TEXT WILL BE INSERTED HERE]", extracted_text.strip())
            try:
                response = self.model.generate_content(
                    full_prompt,
                    generation_config=genai.types.GenerationConfig(
                        response_mime_type="application/json",
                    )
                )
                
                if not response.candidates or not response.candidates[0].content.parts:
                    console.print(f"[warning]No valid response for {pdf_filename}[/warning]")
                    continue
                
                json_response_str = response.candidates[0].content.parts[0].text
            except Exception as e:
                console.print(f"[warning]Gemini API error for {pdf_filename}: {e}[/warning]")
                continue
            
            try:
                data = json.loads(json_response_str)
                for key in ["date", "inward_from", "inward_to", "outward_from", "outward_to", 
                           "visits", "vendor", "inward_charges", "outward_charges"]:
                    if key not in data:
                        data[key] = ""
                
                if not data.get("date"):
                    console.print(f"[warning]No date found in response for {pdf_filename}[/warning]")
                    continue
                
                all_trip_data.append(data)
            except json.JSONDecodeError as e:
                console.print(f"[warning]JSON parse error for {pdf_filename}: {e}[/warning]")
                continue
        
        if not all_trip_data:
            console.print("[warning]No valid trip data extracted[/warning]")
            return False
        
        date_to_row_map = {}
        next_available_row = START_ROW
        excel_rows_data = {}
        
        for item_idx, trip_item in enumerate(all_trip_data):
            date = trip_item.get("date")
            if not date:
                continue
            
            if date in date_to_row_map:
                row_num = date_to_row_map[date]
                existing_row_data = excel_rows_data.get(row_num, {})
                
                for key in ["inward_from", "inward_to", "inward_charges", "outward_from", 
                           "outward_to", "outward_charges", "vendor", "visits"]:
                    current_val = existing_row_data.get(key, "")
                    new_val = trip_item.get(key, "")
                    if not current_val and new_val:
                        existing_row_data[key] = new_val
                
                excel_rows_data[row_num] = existing_row_data
            else:
                row_num = next_available_row
                date_to_row_map[date] = row_num
                
                new_row_data = {
                    "idx": len(excel_rows_data) + 1,
                    "date": date,
                    "inward_from": trip_item.get("inward_from", ""),
                    "inward_to": trip_item.get("inward_to", ""),
                    "outward_from": trip_item.get("outward_from", ""),
                    "outward_to": trip_item.get("outward_to", ""),
                    "visits": trip_item.get("visits", 1),
                    "vendor": trip_item.get("vendor", ""),
                    "inward_charges": str(trip_item.get("inward_charges", "")),
                    "outward_charges": str(trip_item.get("outward_charges", ""))
                }
                excel_rows_data[row_num] = new_row_data
                next_available_row += 1
        
        sorted_excel_items = sorted(excel_rows_data.items(), key=lambda item: item[1]['date'])
        final_excel_data = {}
        current_excel_row = START_ROW
        
        for _, row_data in sorted_excel_items:
            final_excel_data[current_excel_row] = row_data
            final_excel_data[current_excel_row]['idx'] = current_excel_row - START_ROW + 1
            current_excel_row += 1
        
        try:
            for row_num, row_data in final_excel_data.items():
                sheet.cell(row=row_num, column=COLUMN_MAPPING["idx"] + 1, value=row_data.get("idx"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["date"] + 1, value=row_data.get("date"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_from"] + 1, value=row_data.get("inward_from"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_to"] + 1, value=row_data.get("inward_to"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_from"] + 1, value=row_data.get("outward_from"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_to"] + 1, value=row_data.get("outward_to"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["visits"] + 1, value=row_data.get("visits"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["vendor"] + 1, value=row_data.get("vendor"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["inward_charges"] + 1, value=row_data.get("inward_charges"))
                sheet.cell(row=row_num, column=COLUMN_MAPPING["outward_charges"] + 1, value=row_data.get("outward_charges"))
            
            wb.save(output_file)
            return True
            
        except Exception as e:
            console.print(f"[error]Error saving Excel file: {e}[/error]")
            return False

def show_welcome_screen():
    """Display welcome screen with ASCII art"""
    caba_art = r"""                                                     
                                                     
  ,----..     ,---,           ,---,.    ,---,        
 /   /   \   '  .' \        ,'  .'  \  '  .' \       
|   :     : /  ;    '.    ,---.' .' | /  ;    '.     
.   |  ;. /:  :       \   |   |  |: |:  :       \    
.   ; /--` :  |   /\   \  :   :  :  /:  |   /\   \   
;   | ;    |  :  ' ;.   : :   |    ; |  :  ' ;.   :  
|   : |    |  |  ;/  \   \|   :     \|  |  ;/  \   \ 
.   | '___ '  :  | \  \ ,'|   |   . |'  :  | \  \ ,' 
'   ; : .'||  |  '  '--'  '   :  '; ||  |  '  '--'   
'   | '/  :|  :  :        |   |  | ; |  :  :         
|   :    / |  | ,'        |   :   /  |  | ,'         
 \   \ .'  `--''          |   | ,'   `--''           
  `---`                   `----'                     
                                                     """ 
    console.print()
    console.print(Panel(caba_art, title="[primary]Welcome to CABA[/primary]", 
                        border_style="primary", padding=(1, 2)))
    console.print()
    
    subtitle = Text("A professional tool for processing PDF cab booking data", style="info")
    console.print(Panel(subtitle, border_style="secondary", padding=(0, 2)))
    console.print()

def show_completion_screen(processed_count: int, output_file: str):
    """Display completion screen with results and social links"""
    console.print()
    
    success_text = Text(f"✓ Successfully processed {processed_count} PDF file(s) Please verify once !", style="success")
    console.print(Panel(success_text, title="[success]Processing Complete![/success]", 
                        border_style="success", padding=(1, 2)))
    
    console.print(f"\n[info]Output saved to:[/info] [highlight]{output_file}[/highlight]")
    
    social_text = Text()
    social_text.append("Connect with me:\n", style="primary")
    social_text.append("🔗 LinkedIn: ", style="secondary")
    social_text.append("https://www.linkedin.com/in/sudeep-choudhary-103017203/", style="info")
    social_text.append("\n🐦 X : ", style="secondary")
    social_text.append("https://x.com/Sudeep13194447", style="info")
    
    console.print()
    console.print(Panel(social_text, title="[primary]Stay Connected[/primary]", 
                        border_style="primary", padding=(1, 2)))
    
    console.print()
    console.print("[success]Thank you for using CABA! 🚗[/success]")

def get_user_inputs(config: CabaConfig) -> Dict[str, str]:
    """Get user inputs with smart defaults and validation"""
    inputs = {}
    
    console.print("[primary]Configuration[/primary]")
    console.print("=" * 50)
    
    default_pdf = config.get("pdf_dir", "pdf")
    pdf_dir = Prompt.ask(
        "[info]PDF directory path[/info]",
        default=default_pdf,
        show_default=True
    )
    
    while not os.path.isdir(pdf_dir):
        console.print("[error]Directory not found. Please try again.[/error]")
        pdf_dir = Prompt.ask("[info]PDF directory path[/info]")
    
    inputs["pdf_dir"] = pdf_dir
    config.set("pdf_dir", pdf_dir)
    
    template_file = config.get("template_file", "base.xlsx")
    # template_file = Prompt.ask(
    #     "[info]Excel template file[/info] [dim](optional)[/dim]",
    #     default=default_template,
    #     show_default=True
    # )
    
    if template_file and not os.path.isfile(template_file):
        console.print(f"[warning]Template file not found: {template_file}[/warning]")
        if not Confirm.ask("[info]Continue without template?[/info]"):
            template_file = Prompt.ask("[info]Excel template file[/info]")
    
    inputs["template"] = template_file
    config.set("template_file", template_file)
    
    output_file = config.get("output_file", "output.xlsx")
    # output_file = Prompt.ask(
    #     "[info]Output Excel file[/info] [dim](optional)[/dim]",
    #     default=default_output,
    #     show_default=True
    # )
    inputs["output"] = output_file
    config.set("output_file", output_file)
    
    prompt_file = config.get("prompt_file", "prompt.txt")
    # prompt_file = Prompt.ask(
    #     "[info]Prompt file[/info] [dim](optional)[/dim]",
    #     default=default_prompt,
    #     show_default=True
    # )
    
    if prompt_file and not os.path.isfile(prompt_file):
        console.print(f"[warning]Prompt file not found: {prompt_file}[/warning]")
        if not Confirm.ask("[info]Continue without custom prompt?[/info]"):
            prompt_file = Prompt.ask("[info]Prompt file[/info]")
    
    inputs["prompt_file"] = prompt_file
    config.set("prompt_file", prompt_file)
    
    saved_api_key = config.get("gemini_api_key")
    if saved_api_key:
        use_saved = Confirm.ask("[info]Use saved Gemini API key?[/info]")
        if use_saved:
            inputs["gemini_api_key"] = saved_api_key
            return inputs
    
    api_key = Prompt.ask("[info]Gemini API key[/info]", password=True)
    while not api_key:
        console.print("[error]API key is required. Please try again.[/error]")
        api_key = Prompt.ask("[info]Gemini API key[/info]", password=True)
    
    inputs["gemini_api_key"] = api_key
    
    if Confirm.ask("[info]Save API key for future use?[/info]"):
        config.set("gemini_api_key", api_key)
    
    return inputs

def main():
    """Main CLI application"""
    try:
        show_welcome_screen()
        
        config = CabaConfig()
        
        inputs = get_user_inputs(config)
        
        console.print("\n[primary]Processing Configuration[/primary]")
        console.print("=" * 50)
        
        if not inputs.get("template"):
            console.print("[error]Template file is required[/error]")
            return
        
        try:
            processor = PDFProcessor(inputs["gemini_api_key"])
        except Exception:
            console.print("[error]Failed to initialize PDF processor[/error]")
            return
        
        console.print("\n[primary]Starting PDF Processing[/primary]")
        console.print("=" * 50)
        
        def progress_callback(current: int, total: int, filename: str):
            percentage = (current / total) * 100
            console.print(f"[info]Processing [highlight]{filename}[/highlight] ({current}/{total}) - {percentage:.1f}%[/info]")
        
        success = processor.process_pdfs(
            inputs["pdf_dir"],
            inputs["template"],
            inputs["output"],
            inputs.get("prompt_file", "prompt.txt"),
            progress_callback
        )
        
        if success:
            pdf_files = [f for f in os.listdir(inputs["pdf_dir"]) if f.lower().endswith(".pdf")]
            show_completion_screen(len(pdf_files), inputs["output"])
        else:
            console.print("[error]Processing failed. Please check the error messages above.[/error]")
    
    except KeyboardInterrupt:
        console.print("\n[warning]Operation cancelled by user.[/warning]")
    except Exception as e:
        console.print(f"[error]An unexpected error occurred: {e}[/error]")

if __name__ == "__main__":
    main()
